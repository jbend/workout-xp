#!/usr/bin/env python3
"""
One-off dev utility: import a local Essentials spreadsheet into program.json.

The spreadsheet is NOT stored in this repo. Pass the file path explicitly:

  python scripts/import-program.py "C:/path/to/Essentials Program - 5x_Week Spreadsheet.xlsx"

Or set WORKOUT_SOURCE_XLSX.
"""

from __future__ import annotations

import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "packages" / "program-data" / "program.json"
PROGRAM_ID = "jeff-nippard-essentials-5x"
SCHEMA_VERSION = 1


def slugify(value: str) -> str:
    value = value.lower().replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return format_rpe(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def format_rpe(value) -> str:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if float(value).is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, datetime):
        return f"{value.month}-{value.day}"
    return str(value).strip()


def hyperlink(cell) -> str | None:
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    return None


def resolve_source() -> Path:
    if len(sys.argv) > 1:
        return Path(sys.argv[1]).expanduser().resolve()
    env = os.environ.get("WORKOUT_SOURCE_XLSX")
    if env:
        return Path(env).expanduser().resolve()
    raise SystemExit(
        "Spreadsheet path required.\n"
        'Usage: python scripts/import-program.py "path/to/spreadsheet.xlsx"\n'
        "Or set WORKOUT_SOURCE_XLSX."
    )


def build_program(source: Path) -> dict:
    wb = openpyxl.load_workbook(source)
    ws = wb["5x Program"]

    title = ws["B1"].value or "Essentials Program"
    weeks: list[dict] = []
    current_week: dict | None = None
    current_day: dict | None = None

    for row in ws.iter_rows(min_row=1):
        marker = cell_text(row[1].value)
        exercise_name = cell_text(row[2].value)
        week_match = re.match(r"^Week\s+(\d+)", marker, re.I)

        if week_match:
            current_week = {
                "week": int(week_match.group(1)),
                "title": marker,
                "days": [],
            }
            weeks.append(current_week)
            current_day = None
            continue

        if not current_week or marker == "Exercise":
            continue

        if marker and not exercise_name:
            day_order = len(current_week["days"]) + 1
            current_day = {
                "id": f"week-{current_week['week']}-{slugify(marker)}-{day_order}",
                "week": current_week["week"],
                "order": day_order,
                "title": marker,
                "isRestDay": bool(re.search(r"rest", marker, re.I)),
                "exercises": [],
            }
            current_week["days"].append(current_day)
            continue

        if marker and exercise_name:
            day_order = len(current_week["days"]) + 1
            current_day = {
                "id": f"week-{current_week['week']}-{slugify(marker)}-{day_order}",
                "week": current_week["week"],
                "order": day_order,
                "title": marker,
                "isRestDay": False,
                "exercises": [],
            }
            current_week["days"].append(current_day)

        if not current_day or not exercise_name:
            continue

        exercise_order = len(current_day["exercises"]) + 1
        substitutions = []
        for sub_cell in (row[9], row[10]):
            name = cell_text(sub_cell.value)
            if not name:
                continue
            sub_url = hyperlink(sub_cell)
            substitutions.append({"name": name, "videoUrl": sub_url})

        video_links = []
        main_url = hyperlink(row[2])
        if main_url:
            video_links.append({"label": exercise_name, "url": main_url})
        for sub in substitutions:
            if sub["videoUrl"]:
                video_links.append({"label": sub["name"], "url": sub["videoUrl"]})

        current_day["exercises"].append(
            {
                "id": f"{slugify(exercise_name)}-{exercise_order}",
                "order": exercise_order,
                "name": exercise_name,
                "warmupSets": cell_text(row[3].value),
                "workingSets": cell_text(row[4].value),
                "reps": cell_text(row[5].value),
                "load": cell_text(row[6].value),
                "rpe": format_rpe(row[7].value) if row[7].value is not None else "",
                "rest": cell_text(row[8].value),
                "substitutions": substitutions,
                "notes": cell_text(row[11].value),
                "videoLinks": video_links,
            }
        )

    return {
        "schemaVersion": SCHEMA_VERSION,
        "programId": PROGRAM_ID,
        "title": title,
        "weeks": weeks,
    }


def main() -> None:
    source = resolve_source()
    if not source.exists():
        raise SystemExit(f"Source spreadsheet not found: {source}")

    program = build_program(source)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(program, indent=2), encoding="utf-8")

    exercises = sum(
        len(day["exercises"])
        for week in program["weeks"]
        for day in week["days"]
        if not day["isRestDay"]
    )
    videos = sum(
        len(exercise["videoLinks"])
        for week in program["weeks"]
        for day in week["days"]
        for exercise in day["exercises"]
    )
    print(f"Wrote {OUTPUT}")
    print(f"  {len(program['weeks'])} weeks, {exercises} exercises, {videos} video links")


if __name__ == "__main__":
    main()
